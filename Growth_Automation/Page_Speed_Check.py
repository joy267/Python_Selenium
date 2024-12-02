import random
import time

import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

s = Service("C:\\Projects\\Selenium_Automation\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe")
driver = webdriver.Chrome(service=s)

driver.maximize_window()
driver.implicitly_wait(10)

base_url = "https://pagespeed.web.dev/"

# Mobile Page Speed Score Check


#  1 Iteration

print("First iteration is starting...")
time.sleep(1)

driver.get(base_url)

driver.find_element(By.XPATH, "//button/span[text()='Ok, Got it.']").click()
keyword_1 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"
driver.find_element(By.XPATH, "//input[@id='i4']").send_keys(keyword_1)
driver.find_element(By.XPATH, "//span[text()='Analyze']").click()

wait = WebDriverWait(driver, 55)
mobile_score = wait.until(
    expected_conditions.visibility_of_element_located((By.XPATH, "//div[@class='lh-exp-gauge__svg-wrapper']")))

mobile_page_speed_1 = driver.find_element(By.CLASS_NAME, "lh-exp-gauge__percentage").text

time.sleep(1)
driver.quit()

print("First iteration has completed")
time.sleep(1)

# 2 Iteration

print("Second iteration is starting...")
time.sleep(1)

s = Service("C:\\Projects\\Selenium_Automation\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe")
driver = webdriver.Chrome(service=s)
driver.maximize_window()
driver.implicitly_wait(10)

driver.get(base_url)

driver.find_element(By.XPATH, "//button/span[text()='Ok, Got it.']").click()
keyword_2 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"
driver.find_element(By.XPATH, "//input[@id='i4']").send_keys(keyword_2)
driver.find_element(By.XPATH, "//span[text()='Analyze']").click()

wait = WebDriverWait(driver, 55)
mobile_score = wait.until(
    expected_conditions.visibility_of_element_located((By.XPATH, "//div[@class='lh-exp-gauge__svg-wrapper']")))

mobile_page_speed_2 = driver.find_element(By.CLASS_NAME, "lh-exp-gauge__percentage").text

time.sleep(1)
driver.quit()

print("Second iteration has completed")
time.sleep(1)

# 3 Iteration

print("Third iteration is starting...")
time.sleep(1)

s = Service("C:\\Projects\\Selenium_Automation\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe")
driver = webdriver.Chrome(service=s)
driver.maximize_window()
driver.implicitly_wait(10)

driver.get(base_url)

driver.find_element(By.XPATH, "//button/span[text()='Ok, Got it.']").click()
keyword_3 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"
driver.find_element(By.XPATH, "//input[@id='i4']").send_keys(keyword_3)
driver.find_element(By.XPATH, "//span[text()='Analyze']").click()

wait = WebDriverWait(driver, 55)
mobile_score = wait.until(
    expected_conditions.visibility_of_element_located((By.XPATH, "//div[@class='lh-exp-gauge__svg-wrapper']")))

mobile_page_speed_3 = driver.find_element(By.CLASS_NAME, "lh-exp-gauge__percentage").text

time.sleep(1)
driver.quit()

print("Third iteration has completed")
time.sleep(1)

# 4 Iteration

print("Fourth iteration is starting...")
time.sleep(1)

s = Service("C:\\Projects\\Selenium_Automation\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe")
driver = webdriver.Chrome(service=s)
driver.maximize_window()
driver.implicitly_wait(10)

driver.get(base_url)

driver.find_element(By.XPATH, "//button/span[text()='Ok, Got it.']").click()
keyword_4 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"
driver.find_element(By.XPATH, "//input[@id='i4']").send_keys(keyword_4)
driver.find_element(By.XPATH, "//span[text()='Analyze']").click()

wait = WebDriverWait(driver, 55)
mobile_score = wait.until(
    expected_conditions.visibility_of_element_located((By.XPATH, "//div[@class='lh-exp-gauge__svg-wrapper']")))

mobile_page_speed_4 = driver.find_element(By.CLASS_NAME, "lh-exp-gauge__percentage").text

time.sleep(1)
driver.quit()

print("Fourth iteration has completed")
time.sleep(1)

# 5 Iteration

print("Fifth iteration is starting...")
time.sleep(1)

s = Service("C:\\Projects\\Selenium_Automation\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe")
driver = webdriver.Chrome(service=s)
driver.maximize_window()
driver.implicitly_wait(10)

driver.get(base_url)

driver.find_element(By.XPATH, "//button/span[text()='Ok, Got it.']").click()
keyword_5 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"
driver.find_element(By.XPATH, "//input[@id='i4']").send_keys(keyword_5)
driver.find_element(By.XPATH, "//span[text()='Analyze']").click()

wait = WebDriverWait(driver, 55)
mobile_score = wait.until(
    expected_conditions.visibility_of_element_located((By.XPATH, "//div[@class='lh-exp-gauge__svg-wrapper']")))

mobile_page_speed_5 = driver.find_element(By.CLASS_NAME, "lh-exp-gauge__percentage").text

time.sleep(1)
driver.quit()

print("Fifth iteration has completed")
time.sleep(1)


# All iterations have done
print("All iterations have done.")


# Create a new Excel

print("Generating Excel file...")

df = pd.DataFrame()
df.to_excel('Page_score_analysis.xlsx', index=False)


# Open Excel

workbook = openpyxl.load_workbook("Page_score_analysis.xlsx")
sheet = workbook.active

sheet.cell(row=1, column=1).value = 'Iterations'
sheet.cell(row=1, column=2).value = 'Test Keywords'
sheet.cell(row=1, column=3).value = 'Score'

sheet.cell(row=2, column=1).value = '1'
sheet.cell(row=2, column=2).value = keyword_1
sheet.cell(row=2, column=3).value = mobile_page_speed_1

sheet.cell(row=3, column=1).value = '2'
sheet.cell(row=3, column=2).value = keyword_2
sheet.cell(row=3, column=3).value = mobile_page_speed_2

sheet.cell(row=4, column=1).value = '3'
sheet.cell(row=4, column=2).value = keyword_3
sheet.cell(row=4, column=3).value = mobile_page_speed_3

sheet.cell(row=5, column=1).value = '4'
sheet.cell(row=5, column=2).value = keyword_4
sheet.cell(row=5, column=3).value = mobile_page_speed_4

sheet.cell(row=6, column=1).value = '5'
sheet.cell(row=6, column=2).value = keyword_5
sheet.cell(row=6, column=3).value = mobile_page_speed_5

workbook.save("Page_score_analysis.xlsx")

print("Excel file has generated successfully.")

driver.quit()
