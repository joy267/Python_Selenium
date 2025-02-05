import random
import time

import openpyxl
import pandas as pd

from Page_Objects.page_speed import Page_Speed
from datetime import datetime

current_date_time = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")

page_speed = Page_Speed()

Test_Page_URL_1 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"

Test_Page_URL_2 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"

Test_Page_URL_3 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"

Test_Page_URL_4 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"

Test_Page_URL_5 = f"https://staging-ppcbingo.kinsta.cloud/?keyword=page_speed_score_analysis_{random.randint(0, 10000)}"

# Mobile Page Speed Score Check


# 1 Iteration

print("First iteration is starting...")

time.sleep(1)

page_speed.open_webdriver()

page_speed.maximize_window()

page_speed.implicit_wait(10)

page_speed.open_pagespeed_insights("https://pagespeed.web.dev/")

page_speed.accept_cookies()

page_speed.delete_all_cache()

keyword_1 = Test_Page_URL_1

page_speed.test_page_url(keyword_1)

time.sleep(1)

page_speed.click_analyze_button()

time.sleep(1)

page_speed.explicit_wait(90)

first_iteration_performance = page_speed.get_mobile_site_performance()

first_iteration_accessibility = page_speed.get_mobile_accessibility()

first_iteration_best_practices = page_speed.get_mobile_best_practices()

first_iteration_seo = page_speed.get_mobile_seo()

first_iteration_first_contentful_paint = page_speed.get_mobile_first_contentful_paint()

first_iteration_largest_contentful_paint = page_speed.get_mobile_largest_contentful_paint()

first_iteration_total_blocking_time = page_speed.get_mobile_total_blocking_time()

first_iteration_cumulative_layout_shift = page_speed.get_mobile_cumulative_layout_shift()

first_iteration_speed_index = page_speed.get_mobile_speed_index()

time.sleep(1)

page_speed.close_window()

print("First iteration has completed")

time.sleep(1)

print("First iteration --> Page Performance Score :", first_iteration_performance)

time.sleep(1)


# # 2 Iteration

print("Second iteration is starting...")

time.sleep(1)

page_speed.open_webdriver()

page_speed.maximize_window()

page_speed.implicit_wait(10)

page_speed.open_pagespeed_insights("https://pagespeed.web.dev/")

page_speed.accept_cookies()

page_speed.delete_all_cache()

keyword_2 = Test_Page_URL_2

page_speed.test_page_url(keyword_2)

page_speed.click_analyze_button()

time.sleep(1)

page_speed.explicit_wait(90)

second_iteration_performance = page_speed.get_mobile_site_performance()

second_iteration_accessibility = page_speed.get_mobile_accessibility()

second_iteration_best_practices = page_speed.get_mobile_best_practices()

second_iteration_seo = page_speed.get_mobile_seo()

second_iteration_first_contentful_paint = page_speed.get_mobile_first_contentful_paint()

second_iteration_largest_contentful_paint = page_speed.get_mobile_largest_contentful_paint()

second_iteration_total_blocking_time = page_speed.get_mobile_total_blocking_time()

second_iteration_cumulative_layout_shift = page_speed.get_mobile_cumulative_layout_shift()

second_iteration_speed_index = page_speed.get_mobile_speed_index()

time.sleep(1)

page_speed.close_window()

print("Second iteration has completed")

time.sleep(1)

print("Second iteration --> Page Performance Score :", second_iteration_performance)

time.sleep(1)


# # 3 Iteration

print("Third iteration is starting...")

time.sleep(1)

page_speed.open_webdriver()

page_speed.maximize_window()

page_speed.implicit_wait(10)

page_speed.open_pagespeed_insights("https://pagespeed.web.dev/")

page_speed.accept_cookies()

page_speed.delete_all_cache()

keyword_3 = Test_Page_URL_3

page_speed.test_page_url(keyword_3)

page_speed.click_analyze_button()

time.sleep(1)

page_speed.explicit_wait(90)

third_iteration_performance = page_speed.get_mobile_site_performance()

third_iteration_accessibility = page_speed.get_mobile_accessibility()

third_iteration_best_practices = page_speed.get_mobile_best_practices()

third_iteration_seo = page_speed.get_mobile_seo()

third_iteration_first_contentful_paint = page_speed.get_mobile_first_contentful_paint()

third_iteration_largest_contentful_paint = page_speed.get_mobile_largest_contentful_paint()

third_iteration_total_blocking_time = page_speed.get_mobile_total_blocking_time()

third_iteration_cumulative_layout_shift = page_speed.get_mobile_cumulative_layout_shift()

third_iteration_speed_index = page_speed.get_mobile_speed_index()

time.sleep(1)

page_speed.close_window()

print("Third iteration has completed")

time.sleep(1)

print("Third iteration --> Page Performance Score :", third_iteration_performance)

time.sleep(1)


# # 4 Iteration

print("Fourth iteration is starting...")

time.sleep(1)

page_speed.open_webdriver()

page_speed.maximize_window()

page_speed.implicit_wait(10)

page_speed.open_pagespeed_insights("https://pagespeed.web.dev/")

page_speed.accept_cookies()

page_speed.delete_all_cache()

keyword_4 = Test_Page_URL_4

page_speed.test_page_url(keyword_4)

page_speed.click_analyze_button()

time.sleep(1)

page_speed.explicit_wait(90)

fourth_iteration_performance = page_speed.get_mobile_site_performance()

fourth_iteration_accessibility = page_speed.get_mobile_accessibility()

fourth_iteration_best_practices = page_speed.get_mobile_best_practices()

fourth_iteration_seo = page_speed.get_mobile_seo()

fourth_iteration_first_contentful_paint = page_speed.get_mobile_first_contentful_paint()

fourth_iteration_largest_contentful_paint = page_speed.get_mobile_largest_contentful_paint()

fourth_iteration_total_blocking_time = page_speed.get_mobile_total_blocking_time()

fourth_iteration_cumulative_layout_shift = page_speed.get_mobile_cumulative_layout_shift()

fourth_iteration_speed_index = page_speed.get_mobile_speed_index()

time.sleep(1)

page_speed.close_window()

print("Fourth iteration has completed")

time.sleep(1)

print("Fourth iteration --> Page Performance Score :", fourth_iteration_performance)

time.sleep(1)


# # 5 Iteration

print("Fifth iteration is starting...")

time.sleep(1)

page_speed.open_webdriver()

page_speed.maximize_window()

page_speed.implicit_wait(10)

page_speed.open_pagespeed_insights("https://pagespeed.web.dev/")

page_speed.accept_cookies()

page_speed.delete_all_cache()

keyword_5 = Test_Page_URL_5

page_speed.test_page_url(keyword_5)

page_speed.click_analyze_button()

time.sleep(1)

page_speed.explicit_wait(90)

fifth_iteration_performance = page_speed.get_mobile_site_performance()

fifth_iteration_accessibility = page_speed.get_mobile_accessibility()

fifth_iteration_best_practices = page_speed.get_mobile_best_practices()

fifth_iteration_seo = page_speed.get_mobile_seo()

fifth_iteration_first_contentful_paint = page_speed.get_mobile_first_contentful_paint()

fifth_iteration_largest_contentful_paint = page_speed.get_mobile_largest_contentful_paint()

fifth_iteration_total_blocking_time = page_speed.get_mobile_total_blocking_time()

fifth_iteration_cumulative_layout_shift = page_speed.get_mobile_cumulative_layout_shift()

fifth_iteration_speed_index = page_speed.get_mobile_speed_index()

time.sleep(1)

page_speed.close_window()

print("Fifth iteration has completed")

time.sleep(1)

print("Fifth iteration --> Page Performance Score :", fifth_iteration_performance)

time.sleep(1)


# # All iterations have done

print("All iterations have done.")


# # Import Scores in the sheet

# # Create a new Excel

print("Generating Excel file...")

sheet_name = f"Page_Performance_Score_{current_date_time}.xlsx"

df = pd.DataFrame()
df.to_excel(sheet_name, index=False)

time.sleep(1)


# # Open Excel

workbook = openpyxl.load_workbook(sheet_name)
sheet = workbook.active

# # Create rows and columns
sheet.cell(row=1, column=1).value = 'Iterations'
sheet.cell(row=1, column=2).value = 'Test Keywords'
sheet.cell(row=1, column=3).value = 'Performance'
sheet.cell(row=1, column=4).value = 'Accessibility'
sheet.cell(row=1, column=5).value = 'Best Practices'
sheet.cell(row=1, column=6).value = 'SEO'
sheet.cell(row=1, column=7).value = 'First Contentful Paint (FCP)'
sheet.cell(row=1, column=8).value = 'Largest Contentful Paint (LCP)'
sheet.cell(row=1, column=9).value = 'Total Blocking Time (TBT)'
sheet.cell(row=1, column=10).value = 'Cumulative Layout Shift (CLS)'
sheet.cell(row=1, column=11).value = 'Speed Index (SI)'


# # First Interation Data

print("First interation data is importing ...")

sheet.cell(row=2, column=1).value = '1'
time.sleep(1)
sheet.cell(row=2, column=2).value = keyword_1
time.sleep(1)
sheet.cell(row=2, column=3).value = first_iteration_performance
time.sleep(1)
sheet.cell(row=2, column=4).value = first_iteration_accessibility
time.sleep(1)
sheet.cell(row=2, column=5).value = first_iteration_best_practices
time.sleep(1)
sheet.cell(row=2, column=6).value = first_iteration_seo
time.sleep(1)
sheet.cell(row=2, column=7).value = first_iteration_first_contentful_paint
time.sleep(1)
sheet.cell(row=2, column=8).value = first_iteration_largest_contentful_paint
time.sleep(1)
sheet.cell(row=2, column=9).value = first_iteration_total_blocking_time
time.sleep(1)
sheet.cell(row=2, column=10).value = first_iteration_cumulative_layout_shift
time.sleep(1)
sheet.cell(row=2, column=11).value = first_iteration_speed_index
time.sleep(1)

workbook.save(sheet_name)

print("First interation data has imported.")


# # Second Interation Data

print("Second interation data is importing ...")

sheet.cell(row=3, column=1).value = '2'
time.sleep(1)
sheet.cell(row=3, column=2).value = keyword_2
time.sleep(1)
sheet.cell(row=3, column=3).value = second_iteration_performance
time.sleep(1)
sheet.cell(row=3, column=4).value = second_iteration_accessibility
time.sleep(1)
sheet.cell(row=3, column=5).value = second_iteration_best_practices
time.sleep(1)
sheet.cell(row=3, column=6).value = second_iteration_seo
time.sleep(1)
sheet.cell(row=3, column=7).value = second_iteration_first_contentful_paint
time.sleep(1)
sheet.cell(row=3, column=8).value = second_iteration_largest_contentful_paint
time.sleep(1)
sheet.cell(row=3, column=9).value = second_iteration_total_blocking_time
time.sleep(1)
sheet.cell(row=3, column=10).value = second_iteration_cumulative_layout_shift
time.sleep(1)
sheet.cell(row=3, column=11).value = second_iteration_speed_index
time.sleep(1)

workbook.save(sheet_name)

print("Second interation data has imported.")


# # Third Interation Data

print("Third interation data is importing ...")

sheet.cell(row=4, column=1).value = '3'
time.sleep(1)
sheet.cell(row=4, column=2).value = keyword_3
time.sleep(1)
sheet.cell(row=4, column=3).value = third_iteration_performance
time.sleep(1)
sheet.cell(row=4, column=4).value = third_iteration_accessibility
time.sleep(1)
sheet.cell(row=4, column=5).value = third_iteration_best_practices
time.sleep(1)
sheet.cell(row=4, column=6).value = third_iteration_seo
time.sleep(1)
sheet.cell(row=4, column=7).value = third_iteration_first_contentful_paint
time.sleep(1)
sheet.cell(row=4, column=8).value = third_iteration_largest_contentful_paint
time.sleep(1)
sheet.cell(row=4, column=9).value = third_iteration_total_blocking_time
time.sleep(1)
sheet.cell(row=4, column=10).value = third_iteration_cumulative_layout_shift
time.sleep(1)
sheet.cell(row=4, column=11).value = third_iteration_speed_index
time.sleep(1)

workbook.save(sheet_name)

print("Third interation data has imported.")


# # Fourth Interation Data

print("Fourth interation data is importing ...")

sheet.cell(row=5, column=1).value = '4'
time.sleep(1)
sheet.cell(row=5, column=2).value = keyword_4
time.sleep(1)
sheet.cell(row=5, column=3).value = fourth_iteration_performance
time.sleep(1)
sheet.cell(row=5, column=4).value = fourth_iteration_accessibility
time.sleep(1)
sheet.cell(row=5, column=5).value = fourth_iteration_best_practices
time.sleep(1)
sheet.cell(row=5, column=6).value = fourth_iteration_seo
time.sleep(1)
sheet.cell(row=5, column=7).value = fourth_iteration_first_contentful_paint
time.sleep(1)
sheet.cell(row=5, column=8).value = fourth_iteration_largest_contentful_paint
time.sleep(1)
sheet.cell(row=5, column=9).value = fourth_iteration_total_blocking_time
time.sleep(1)
sheet.cell(row=5, column=10).value = fourth_iteration_cumulative_layout_shift
time.sleep(1)
sheet.cell(row=5, column=11).value = fourth_iteration_speed_index
time.sleep(1)

workbook.save(sheet_name)

print("Fourth interation data has imported.")


# # Fifth Interation Data

print("Fifth interation data is importing ...")

sheet.cell(row=6, column=1).value = '5'
time.sleep(1)
sheet.cell(row=6, column=2).value = keyword_5
time.sleep(1)
sheet.cell(row=6, column=3).value = fifth_iteration_performance
time.sleep(1)
sheet.cell(row=6, column=4).value = fifth_iteration_accessibility
time.sleep(1)
sheet.cell(row=6, column=5).value = fifth_iteration_best_practices
time.sleep(1)
sheet.cell(row=6, column=6).value = fifth_iteration_seo
time.sleep(1)
sheet.cell(row=6, column=7).value = fifth_iteration_first_contentful_paint
time.sleep(1)
sheet.cell(row=6, column=8).value = fifth_iteration_largest_contentful_paint
time.sleep(1)
sheet.cell(row=6, column=9).value = fifth_iteration_total_blocking_time
time.sleep(1)
sheet.cell(row=6, column=10).value = fifth_iteration_cumulative_layout_shift
time.sleep(1)
sheet.cell(row=6, column=11).value = fifth_iteration_speed_index
time.sleep(1)

workbook.save(sheet_name)

print("Fifth interation data has imported.")

time.sleep(1)

page_speed.close_window()

print("Excel file has generated successfully.")
