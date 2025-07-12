import random
import time
from datetime import datetime

import openpyxl
import pandas as pd

import Page_Objects.console_logs
from Page_Objects.console_logs import Console_Logs

current_date_time = datetime.now().strftime("%d-%m-%Y_%H:%M")

console_logs = Console_Logs()

Page_Objects.console_logs.configure_chrome_options_for_logging()

console_logs.open_webdriver()

time.sleep(1)

console_logs.maximize_window()

time.sleep(1)

console_logs.implicit_wait(10)

time.sleep(1)

console_logs.open_website(
    f"https://staging-ppcbingo.kinsta.cloud/?keyword=test_{random.randint(0, 10000)}_&enableconsole")

time.sleep(1)

console_logs.scroll_website()

time.sleep(2)

logs = console_logs.fetch_console_logs()

level = []
for entry in logs:
    level.append({
         'level': entry['level'],
         'message': entry['message'],
         'source': entry['source'],
         'timestamp': entry['timestamp']
    })


# Generating Excel File

print("Generating Excel file...")

sheet_name = "Console_Logs.xlsx"

try:
    workbook = openpyxl.load_workbook(sheet_name)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

sheet = workbook.active

time.sleep(1)

headers = ['Level', 'Message', 'Source', 'Timestamp']
for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num).value = header

time.sleep(1)

for row_num, record in enumerate(level, start=2):
    sheet.cell(row=row_num, column=1).value = record['level']
    sheet.cell(row=row_num, column=2).value = record['message']
    sheet.cell(row=row_num, column=3).value = record['source']
    sheet.cell(row=row_num, column=4).value = record['timestamp']


workbook.save(sheet_name)

time.sleep(1)

console_logs.close_browser()

print("Excel file has been generated.")
